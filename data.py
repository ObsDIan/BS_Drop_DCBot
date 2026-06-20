# data.py

import csv
import re
from pathlib import Path

from opencc import OpenCC
from openpyxl import load_workbook


# =========================
# 檔案設定
# =========================

INPUT_EXCEL = Path("掉落列表三倍结算.zh-TW.xlsx")
OUTPUT_CSV = Path("BotData.csv")

# 簡體轉繁體，偏台灣用語
cc = OpenCC("s2twp")


# =========================
# 原始 Excel 欄位對應
# =========================
# key：原始 Excel 欄位名稱
# value：(來源類型, 掉落分類)

SOURCE_MAP = {
    "材料掉落": ("副本掉落", "材料"),
    "裝備": ("副本掉落", "裝備"),
    "每日任務獎勵": ("每日任務", "任務獎勵"),
    "斬首任務獎勵": ("斬首任務", "任務獎勵"),
}

OUTPUT_HEADERS = [
    "副本名稱",
    "來源類型",
    "掉落分類",
    "掉落物名稱",
]


# =========================
# 工具函式
# =========================

def to_traditional(value) -> str:
    """
    統一文字：
    - None 轉空字串
    - 去除前後空白
    - 簡體轉繁體
    """
    if value is None:
        return ""

    text = str(value).strip()

    if not text:
        return ""

    return cc.convert(text)


def split_items(value):
    """
    將一格內的多個掉落物拆開。
    支援：
    - 頓號
    - 逗號
    - 斜線
    - 換行
    - 分號
    """
    text = to_traditional(value)

    if not text or text in {"-", "無", "/", "／"}:
        return []

    text = (
        text.replace("\n", "、")
            .replace("，", "、")
            .replace(",", "、")
            .replace("/", "、")
            .replace("／", "、")
            .replace(";", "、")
            .replace("；", "、")
    )

    items = [
        item.strip()
        for item in re.split(r"[、]+", text)
        if item.strip()
    ]

    # 去重，但保留原順序
    return list(dict.fromkeys(items))


def find_header_row(ws, header_name="副本名稱") -> int:
    """
    自動尋找標題列。
    預設在前 10 列內尋找「副本名稱」。
    """
    target = to_traditional(header_name)

    for row in ws.iter_rows(min_row=1, max_row=10):
        values = [to_traditional(cell.value) for cell in row]

        if target in values:
            return row[0].row

    raise RuntimeError(f"找不到「{header_name}」欄位，請確認 Excel 格式。")


def read_headers(ws, header_row: int) -> dict:
    """
    讀取標題列，建立：
    欄位名稱 -> 欄位位置
    """
    headers = {}

    for cell in ws[header_row]:
        header = to_traditional(cell.value)

        if header:
            headers[header] = cell.column

    return headers


# =========================
# 主程式
# =========================

def main():
    if not INPUT_EXCEL.exists():
        raise FileNotFoundError(f"找不到 Excel 檔案：{INPUT_EXCEL}")

    wb = load_workbook(INPUT_EXCEL, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws)
    headers = read_headers(ws, header_row)

    dungeon_col = headers.get("副本名稱")

    if not dungeon_col:
        raise RuntimeError("找不到「副本名稱」欄位。")

    rows = []
    seen = set()

    for row_index in range(header_row + 1, ws.max_row + 1):
        dungeon = to_traditional(ws.cell(row=row_index, column=dungeon_col).value)

        if not dungeon:
            continue

        for source_column, source_info in SOURCE_MAP.items():
            source_column = to_traditional(source_column)

            if source_column not in headers:
                continue

            source_type, category = source_info
            source_type = to_traditional(source_type)
            category = to_traditional(category)

            col_index = headers[source_column]
            items = split_items(ws.cell(row=row_index, column=col_index).value)

            for item in items:
                key = (
                    dungeon,
                    source_type,
                    category,
                    item,
                )

                if key in seen:
                    continue

                seen.add(key)

                rows.append({
                    "副本名稱": dungeon,
                    "來源類型": source_type,
                    "掉落分類": category,
                    "掉落物名稱": item,
                    "數量": "",
                    "備註": "",
                })

    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    print(f"已輸出：{OUTPUT_CSV}")
    print(f"資料筆數：{len(rows)}")


if __name__ == "__main__":
    main()